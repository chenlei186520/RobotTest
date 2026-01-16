from flask import Flask, render_template, request, jsonify, send_file
from io import BytesIO
from datetime import datetime
import config
import test_data
from test_data import TABdisplayconfig
import apis
import time
import threading

# 导入日志配置（必须在其他导入之前，以便重定向print）
import logger_config

app = Flask(__name__)

# SSH连接池（用于相机测试，按ssh_host存储）
# 格式：{ssh_host: {'ssh': ssh_client, 'lock': threading.Lock()}}
camera_ssh_connections = {}
ssh_lock = threading.Lock()  # 保护连接池的锁

def get_categories_by_vehicle(vehicle_model):
    """根据车型获取对应的测试类别列表（优先从CSV文件读取）"""
    # 首先尝试从CSV文件读取配置
    tab_config = TABdisplayconfig.get_tabs_by_device_model(vehicle_model)
    
    if tab_config and 'tabs' in tab_config:
        tab_ids = tab_config['tabs']
        # 根据配置的TAB ID列表，从所有类别中筛选
        categories = [cat for cat in test_data.TEST_CATEGORIES if cat['id'] in tab_ids]
        # 保持配置中的顺序
        category_dict = {cat['id']: cat for cat in categories}
        result = [category_dict[tab_id] for tab_id in tab_ids if tab_id in category_dict]
        if result:
            print(f"[TAB配置] 从CSV文件加载 {vehicle_model} 的TAB配置: {[cat['id'] for cat in result]}")
            return result
    
    # 如果CSV中没有找到，回退到config.py中的配置
    if vehicle_model in config.VEHICLE_TAB_CONFIG:
        tab_ids = config.VEHICLE_TAB_CONFIG[vehicle_model]
        # 根据配置的TAB ID列表，从所有类别中筛选
        categories = [cat for cat in test_data.TEST_CATEGORIES if cat['id'] in tab_ids]
        # 保持配置中的顺序
        category_dict = {cat['id']: cat for cat in categories}
        result = [category_dict[tab_id] for tab_id in tab_ids if tab_id in category_dict]
        if result:
            print(f"[TAB配置] 从config.py加载 {vehicle_model} 的TAB配置: {[cat['id'] for cat in result]}")
            return result
    
    # 如果都没有配置，返回所有类别
    print(f"[TAB配置] 未找到 {vehicle_model} 的配置，返回所有TAB")
    return test_data.TEST_CATEGORIES

@app.route('/api/vehicle_types')
def get_vehicle_types():
    """获取设备类型列表（X060、X080、X100、X150等）"""
    try:
        import csv
        import os
        import re
        
        csv_path = os.path.join(os.path.dirname(__file__), 'test_data', 'TABdisplay_data.csv')
        vehicle_types = set()
        vehicle_models = []
        
        # 读取CSV文件
        encodings = ['utf-8-sig', 'utf-8', 'gbk', 'gb2312']
        for encoding in encodings:
            try:
                with open(csv_path, 'r', encoding=encoding) as f:
                    reader = csv.reader(f)
                    rows = list(reader)
                    
                    # 跳过表头
                    for row in rows[1:]:
                        if len(row) < 1:
                            continue
                        device_model = row[0].strip()
                        if not device_model:
                            continue
                        
                        vehicle_models.append(device_model)
                        
                        # 提取设备类型（X060、X080、X100、X150等）
                        # 匹配 X-060、X-080、X-100、X-150 等格式
                        match = re.match(r'X-?(\d+)', device_model)
                        if match:
                            type_code = 'X' + match.group(1)
                            vehicle_types.add(type_code)
                    
                    break
            except Exception as e:
                print(f"[设备类型] 使用 {encoding} 编码读取CSV失败: {e}")
                continue
        
        # 排序设备类型
        sorted_types = sorted(list(vehicle_types))
        
        return jsonify({
            "status": "success",
            "vehicle_types": sorted_types,
            "all_models": vehicle_models
        })
    except Exception as e:
        print(f"[设备类型] 获取设备类型失败: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/vehicle_models')
def get_vehicle_models():
    """根据设备类型获取具体型号列表"""
    try:
        import csv
        import os
        import re
        
        vehicle_type = request.args.get('type', '')  # 如 X060、X080等
        
        if not vehicle_type:
            return jsonify({"status": "error", "message": "缺少设备类型参数"}), 400
        
        csv_path = os.path.join(os.path.dirname(__file__), 'test_data', 'TABdisplay_data.csv')
        vehicle_models = []
        
        # 读取CSV文件
        encodings = ['utf-8-sig', 'utf-8', 'gbk', 'gb2312']
        for encoding in encodings:
            try:
                with open(csv_path, 'r', encoding=encoding) as f:
                    reader = csv.reader(f)
                    rows = list(reader)
                    
                    # 跳过表头
                    for row in rows[1:]:
                        if len(row) < 1:
                            continue
                        device_model = row[0].strip()
                        if not device_model:
                            continue
                        
                        # 检查是否以该设备类型开头
                        # 支持 X-060 和 X060 两种格式
                        type_pattern = vehicle_type.replace('X', 'X-?')
                        if re.match(f'^{type_pattern}', device_model):
                            vehicle_models.append(device_model)
                    
                    break
            except Exception as e:
                print(f"[设备型号] 使用 {encoding} 编码读取CSV失败: {e}")
                continue
        
        # 去重并排序
        vehicle_models = sorted(list(set(vehicle_models)))
        
        return jsonify({
            "status": "success",
            "vehicle_models": vehicle_models
        })
    except Exception as e:
        print(f"[设备型号] 获取设备型号失败: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/check_vehicle_id', methods=['POST'])
def check_vehicle_id():
    """检测车辆ID是否匹配（通过SSH获取HOSTNAME并与输入的车辆ID比较）"""
    data = request.json
    vehicle_id = data.get('vehicle_id', '').strip()
    ssh_host = data.get('ssh_host', '').strip()
    ssh_user = data.get('ssh_user')
    ssh_password = data.get('ssh_password')
    
    if not vehicle_id:
        return jsonify({"status": "error", "message": "未提供车辆ID"}), 400
    
    if not ssh_host:
        return jsonify({"status": "error", "message": "未提供车辆IP地址"}), 400
    
    ssh_user = ssh_user or config.SSH_USER
    ssh_password = ssh_password or config.SSH_PASSWORD
    
    try:
        import paramiko
        import time
        import re
        
        # 1. 建立SSH连接
        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        ssh.connect(
            hostname=ssh_host,
            port=22,
            username=ssh_user,
            password=ssh_password,
            timeout=10,
            allow_agent=False,
            look_for_keys=False
        )
        print(f"[车辆ID检测] SSH连接成功，开始获取HOSTNAME...")
        
        # 2. 创建交互式Shell通道
        channel = ssh.invoke_shell()
        channel.settimeout(10)
        channel.set_combine_stderr(True)
        
        # 3. 等待shell初始化，清空初始输出
        time.sleep(1)
        if channel.recv_ready():
            initial_output = channel.recv(4096).decode('utf-8', errors='ignore')
            print(f"[车辆ID检测] 清空初始shell提示符: {initial_output[:100]}...")
        
        # 4. 发送获取HOSTNAME的命令（参考get_system_info的实现）
        command = "source ~/.bashrc 2>/dev/null; source /etc/profile 2>/dev/null; echo HOSTNAME=$HOSTNAME\n"
        print(f"[车辆ID检测] 发送命令: {command.strip()}")
        channel.send(command)
        time.sleep(2)  # 等待命令执行和输出返回（确保环境变量加载完成）
        
        # 5. 实时读取输出并解析（参考get_system_info的方式）
        output = ""
        start_time = time.time()
        timeout = 5  # 读取超时时间5秒
        
        print(f"[车辆ID检测] 开始接收输出（超时时间{timeout}秒）...")
        
        while time.time() - start_time < timeout:
            if channel.recv_ready():
                # 读取通道数据（按字节读取，避免截断）
                data = channel.recv(4096).decode('utf-8', errors='ignore')
                if data:
                    output += data
                    print(f"[车辆ID检测] 接收到数据: {data.strip()}")
                    # 如果已经获取到HOSTNAME信息，可以提前退出
                    if 'HOSTNAME=' in output and not output.strip().endswith('HOSTNAME='):
                        print(f"[车辆ID检测] 已获取到HOSTNAME信息，提前退出")
                        break
            time.sleep(0.5)  # 每隔0.5秒读取一次
        
        # 6. 关闭通道
        channel.close()
        
        print(f"[车辆ID检测] 原始输出内容:\n{output}")
        
        # 7. 解析输出，提取HOSTNAME（参考get_system_info的解析逻辑）
        hostname = None
        
        if output:
            # 使用正则表达式提取HOSTNAME（匹配格式：HOSTNAME=值）
            hostname_pattern = r'HOSTNAME\s*=\s*([^\s\n\r;]+)'
            match = re.search(hostname_pattern, output)
            if match:
                hostname = match.group(1).strip()
                # 过滤掉未解析的变量（如 $HOSTNAME）
                if not hostname.startswith('$'):
                    print(f"[车辆ID检测] 正则匹配: HOSTNAME={hostname}")
                else:
                    print(f"[车辆ID检测] 警告: HOSTNAME 的值未解析: {hostname}")
                    hostname = None
            
            # 如果正则匹配失败，使用行解析方式
            if not hostname:
                print(f"[车辆ID检测] 正则匹配失败，使用行解析方式")
                lines = output.split('\n')
                for line in lines:
                    line = line.strip()
                    # 提取包含HOSTNAME的行，排除提示符和命令回显
                    if '=' in line and 'HOSTNAME' in line:
                        # 移除可能的提示符前缀（如 [user@host ~]$ 等）
                        if line.startswith('[') and ']' in line:
                            line = line.split(']', 1)[1].strip()
                        # 移除命令回显（如 echo HOSTNAME=...）
                        if line.startswith('echo '):
                            continue
                        if line and '=' in line:
                            key, value = line.split('=', 1)
                            key = key.strip()
                            value = value.strip()
                            # 过滤掉未解析的变量
                            if key == 'HOSTNAME' and not value.startswith('$'):
                                hostname = value
                                print(f"[车辆ID检测] 行解析: HOSTNAME={hostname}")
                                break
        
        # 8. 关闭SSH连接
        ssh.close()
        
        print(f"[车辆ID检测] 输入的车辆ID: {vehicle_id}, 获取的HOSTNAME: {hostname}")
        print(f"[车辆ID检测] 原始输出内容:\n{output}")
        
        if not hostname:
            print(f"[车辆ID检测] ❌ 无法解析HOSTNAME，原始输出长度: {len(output)}")
            return jsonify({
                "status": "error",
                "message": f"无法获取设备HOSTNAME，请检查SSH连接和网络设置。原始输出: {output[:200]}"
            }), 500
        
        # 8. 比较车辆ID和HOSTNAME
        print(f"[车辆ID检测] 比较: 车辆ID='{vehicle_id}' vs HOSTNAME='{hostname}'")
        if vehicle_id == hostname:
            print(f"[车辆ID检测] ✅ 匹配成功")
            return jsonify({
                "status": "success",
                "matched": True,
                "hostname": hostname,
                "message": "车辆ID匹配成功"
            })
        else:
            print(f"[车辆ID检测] ❌ 匹配失败")
            return jsonify({
                "status": "success",
                "matched": False,
                "hostname": hostname,
                "message": "车辆ID与设备不匹配"
            })
            
    except paramiko.AuthenticationException as e:
        print(f"[车辆ID检测] ❌ SSH认证失败: {e}")
        return jsonify({"status": "error", "message": f"SSH认证失败，请检查用户名和密码: {str(e)}"}), 401
    except paramiko.SSHException as e:
        print(f"[车辆ID检测] ❌ SSH连接错误: {e}")
        return jsonify({"status": "error", "message": f"SSH连接错误: {str(e)}"}), 500
    except Exception as e:
        import traceback
        print(f"[车辆ID检测] ❌ 检测失败: {e}")
        print(f"[车辆ID检测] 错误堆栈:\n{traceback.format_exc()}")
        return jsonify({"status": "error", "message": f"检测车辆ID失败: {str(e)}"}), 500

@app.route('/')
def config_page():
    """配置页面（页面A）"""
    return render_template('config.html')

@app.route('/api/devicetest')
def index():
    # 从URL参数获取车型信息
    vehicle_model = request.args.get('vehiclemodel', 'X100')  # 默认X100
    hostname = request.args.get('hostname', '')
    carip = request.args.get('carip', '')
    
    # 根据车型获取对应的测试类别
    categories = get_categories_by_vehicle(vehicle_model)
    
    if not categories:
        categories = test_data.TEST_CATEGORIES
    
    # 获取第一个测试类别的详情
    first_category = categories[0]
    first_test_id = first_category['id']
    first_test_details = test_data.TEST_DETAILS.get(first_test_id)
    
    # 确保 first_test_details 是字典且包含 sections
    if not first_test_details or 'sections' not in first_test_details:
        first_test_details = {'sections': []}
    
    return render_template('index.html', 
                         test_categories=categories, 
                         all_test_details=test_data.TEST_DETAILS,
                         first_test_id=first_test_id,
                         first_test_details=first_test_details,
                         vehicle_model=vehicle_model,
                         hostname=hostname,
                         carip=carip)

@app.route('/api/test_data/<test_id>')
def get_test_data(test_id):
    if test_id in test_data.TEST_DETAILS:
        test_data_obj = test_data.TEST_DETAILS[test_id]
        
        vehicle_model = request.args.get('vehiclemodel', 'X100')
        
        # 如果是相机测试，根据设备型号从CSV配置中获取需要显示的设备
        if test_id == 'camera':
            # 从CSV配置中获取该设备型号需要显示的相机设备
            camera_devices = TABdisplayconfig.get_camera_devices_by_device_model(vehicle_model)
            print(f"[相机测试] 设备型号 {vehicle_model} 需要显示的设备: {camera_devices}")
            # 使用X100的IP映射作为默认值
            ip_map = config.get_camera_ip_map('X100')
            
            # 过滤设备列表并添加IP地址
            filtered_data = {
                "sections": []
            }
            for section in test_data_obj.get('sections', []):
                filtered_section = {
                    "title": section.get('title', ''),
                    "items": []
                }
                for item in section.get('items', []):
                    # 只显示CSV配置中指定的设备
                    if item.get('id') in camera_devices:
                        # 添加IP地址到设备数据中
                        item_with_ip = item.copy()
                        item_with_ip['default_ip'] = ip_map.get(item.get('id'), '192.168.1.1')
                        filtered_section['items'].append(item_with_ip)
                
                if filtered_section['items']:
                    filtered_data['sections'].append(filtered_section)
            
            return jsonify(filtered_data)
        
        # 如果是按键测试，根据车型过滤测试项
        if test_id == 'button':
            allowed_items = config.get_button_test_items(vehicle_model)
            
            # 过滤测试项列表
            filtered_data = {
                "sections": []
            }
            for section in test_data_obj.get('sections', []):
                filtered_section = {
                    "title": section.get('title', ''),
                    "items": []
                }
                for item in section.get('items', []):
                    if item.get('id') in allowed_items:
                        filtered_section['items'].append(item)
                
                if filtered_section['items']:
                    filtered_data['sections'].append(filtered_section)
            
            return jsonify(filtered_data)
        
        return jsonify(test_data_obj)
    return jsonify({"error": "Test not found"}), 404

@app.route('/api/config')
def get_config():
    """获取前端配置"""
    return jsonify({
        "io_check_timeout": config.IO_CHECK_TIMEOUT,
        "io_check_interval": config.IO_CHECK_INTERVAL,
        "command_wait_time": config.COMMAND_WAIT_TIME
    })

@app.route('/api/get_system_info', methods=['POST'])
def get_system_info():
    """通过SSH获取系统信息（PRODUCT_NAME, PRODUCT_NAME_EXTERNAL, HOSTNAME, APP_VERSION）"""
    data = request.json
    ssh_host = data.get('ssh_host')
    ssh_user = data.get('ssh_user')
    ssh_password = data.get('ssh_password')
    
    if not ssh_host:
        return jsonify({"status": "error", "message": "未提供SSH主机地址"}), 400
    
    ssh_user = ssh_user or config.SSH_USER
    ssh_password = ssh_password or config.SSH_PASSWORD
    
    try:
        import paramiko
        import time
        import re
        
        # 1. 建立SSH连接（参考工作代码）
        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        ssh.connect(
            hostname=ssh_host,
            port=22,
            username=ssh_user,
            password=ssh_password,
            timeout=10,
            allow_agent=False,
            look_for_keys=False
        )
        print(f"[系统信息调试] ✅ SSH连接成功，开始获取系统信息...")
        
        # 2. 创建交互式Shell通道（关键：支持持续输出的指令）
        channel = ssh.invoke_shell()
        channel.settimeout(10)  # 通道超时时间
        channel.set_combine_stderr(True)  # 合并标准输出和错误输出
        
        # 3. 等待shell初始化，清空初始输出
        time.sleep(1)
        if channel.recv_ready():
            initial_output = channel.recv(4096).decode('utf-8', errors='ignore')
            print(f"[系统信息调试] 清空初始shell提示符: {initial_output[:100]}...")
        
        # 4. 发送获取环境变量的命令
        # 先尝试source环境变量文件（如果存在），然后获取环境变量
        command = "source ~/.bashrc 2>/dev/null; source /etc/profile 2>/dev/null; echo PRODUCT_NAME=$PRODUCT_NAME; echo PRODUCT_NAME_EXTERNAL=$PRODUCT_NAME_EXTERNAL; echo HOSTNAME=$HOSTNAME; echo APP_VERSION=$APP_VERSION\n"
        print(f"[系统信息调试] 发送命令: {command.strip()}")
        channel.send(command)
        time.sleep(2)  # 等待命令执行和输出返回（确保环境变量加载完成）
        
        # 5. 实时读取输出并解析（参考工作代码的方式）
        output = ""
        start_time = time.time()
        timeout = 5  # 读取超时时间5秒
        
        print(f"[系统信息调试] 开始接收输出（超时时间{timeout}秒）...")
        
        while time.time() - start_time < timeout:
            if channel.recv_ready():
                # 读取通道数据（按字节读取，避免截断）
                data = channel.recv(4096).decode('utf-8', errors='ignore')
                if data:
                    output += data
                    print(f"[系统信息调试] 接收到数据: {data.strip()}")  # 打印原始输出
                    # 如果已经获取到所有需要的信息，可以提前退出
                    if all(key in output for key in ['PRODUCT_NAME=', 'HOSTNAME=', 'APP_VERSION=']):
                        print(f"[系统信息调试] 已获取到所有需要的信息，提前退出")
                        break
            time.sleep(0.5)  # 每隔0.5秒读取一次
        
        # 6. 关闭通道
        channel.close()
        
        print(f"[系统信息调试] 原始输出内容:\n{output}")
        
        # 7. 解析输出，提取环境变量
        system_info = {}
        
        if output:
            # 使用正则表达式提取环境变量（匹配格式：PRODUCT_NAME=值）
            env_patterns = {
                'PRODUCT_NAME': r'PRODUCT_NAME\s*=\s*([^\s\n\r;]+)',
                'PRODUCT_NAME_EXTERNAL': r'PRODUCT_NAME_EXTERNAL\s*=\s*([^\s\n\r;]+)',
                'HOSTNAME': r'HOSTNAME\s*=\s*([^\s\n\r;]+)',
                'APP_VERSION': r'APP_VERSION\s*=\s*([^\s\n\r;]+)'
            }
            
            # 正则匹配
            for key, pattern in env_patterns.items():
                match = re.search(pattern, output)
                if match:
                    value = match.group(1).strip()
                    # 过滤掉未解析的变量（如 $HOSTNAME）
                    if not value.startswith('$'):
                        system_info[key] = value
                        print(f"[系统信息调试] 正则匹配: {key}={value}")
                    else:
                        print(f"[系统信息调试] 警告: {key} 的值未解析: {value}")
            
            # 如果正则匹配失败，使用行解析方式
            if len(system_info) < 3:
                print(f"[系统信息调试] 正则匹配不完整（只匹配到{len(system_info)}个字段），使用行解析方式")
                lines = output.split('\n')
                for line in lines:
                    line = line.strip()
                    # 提取包含环境变量的行，排除提示符和命令回显
                    if '=' in line and any(key in line for key in ['PRODUCT_NAME', 'PRODUCT_NAME_EXTERNAL', 'HOSTNAME', 'APP_VERSION']):
                        # 移除可能的提示符前缀（如 [user@host ~]$ 等）
                        if line.startswith('[') and ']' in line:
                            line = line.split(']', 1)[1].strip()
                        # 移除命令回显（如 echo PRODUCT_NAME=...）
                        if line.startswith('echo '):
                            continue
                        if line and '=' in line:
                            key, value = line.split('=', 1)
                            key = key.strip()
                            value = value.strip()
                            # 过滤掉未解析的变量
                            if key in ['PRODUCT_NAME', 'PRODUCT_NAME_EXTERNAL', 'HOSTNAME', 'APP_VERSION'] and not value.startswith('$'):
                                system_info[key] = value
                                print(f"[系统信息调试] 行解析: {key}={value}")
        
        # 8. 检查是否成功获取到系统信息
        if system_info and len(system_info) >= 3:  # 至少获取到3个关键字段
            ssh.close()
            print(f"[系统信息调试] ✅ 成功获取系统信息: {system_info}")
            
            return jsonify({
                "status": "success",
                "data": {
                    "PRODUCT_NAME": system_info.get('PRODUCT_NAME', ''),
                    "PRODUCT_NAME_EXTERNAL": system_info.get('PRODUCT_NAME_EXTERNAL', ''),
                    "HOSTNAME": system_info.get('HOSTNAME', ''),
                    "APP_VERSION": system_info.get('APP_VERSION', '')
                }
            })
        else:
            # 如果获取失败
            print(f"[系统信息调试] ❌ 获取系统信息失败")
            print(f"[系统信息调试] 解析到的系统信息: {system_info}")
            print(f"[系统信息调试] 原始输出: {output[:500]}...")
            
            ssh.close()
            return jsonify({
                "status": "error", 
                "message": f"无法获取系统信息。请检查环境变量是否正确设置。解析到的信息: {system_info}, 原始输出长度: {len(output) if output else 0}"
            }), 500
            
    except paramiko.AuthenticationException:
        return jsonify({"status": "error", "message": "SSH认证失败"}), 401
    except paramiko.SSHException as e:
        return jsonify({"status": "error", "message": f"SSH连接错误: {str(e)}"}), 500
    except Exception as e:
        return jsonify({"status": "error", "message": f"获取系统信息失败: {str(e)}"}), 500

@app.route('/api/vehicle_tabs')
def get_vehicle_tabs():
    """根据车型获取对应的TAB列表"""
    vehicle_model = request.args.get('vehiclemodel', 'X100')
    categories = get_categories_by_vehicle(vehicle_model)
    return jsonify({
        "vehicle_model": vehicle_model,
        "categories": categories
    })

@app.route('/api/send_command', methods=['POST'])
def send_command():
    """发送指令到串口（根据测试类型路由到对应的API处理器，通过SSH执行）"""
    data = request.json
    item_id = data.get('item_id')
    test_id = data.get('test_id')  # 可选，用于确定使用哪个API处理器
    ssh_host = data.get('ssh_host')  # SSH主机地址（从URL的carip参数获取）
    ssh_user = data.get('ssh_user')  # SSH用户名（可选，默认使用config中的配置）
    ssh_password = data.get('ssh_password')  # SSH密码（可选，默认使用config中的配置）
    
    # 如果没有提供test_id，根据item_id推断（主要用于灯光测试）
    if not test_id:
        if item_id in config.COMMAND_MAP or item_id == 'turn_off_all_lights':
            test_id = 'light'
        else:
            test_id = 'default'
    
    # 获取对应的API处理器
    api_handler = apis.get_api_handler(test_id)
    
    # 调用对应的发送指令方法，传递SSH信息
    # 对于有专门API处理器的测试类型（light、voice、button、touch、display、lift_motor、rotation_motor），直接调用
    # 对于没有专门API处理器的测试类型，使用BaseAPI（command_map为None，表示允许无命令的测试项）
    if test_id in ['light', 'voice', 'button', 'touch', 'display']:
        # 这些测试类型使用专门的API处理器
        result = api_handler.send_command(item_id, ssh_host, ssh_user, ssh_password)
    elif test_id == 'lift_motor':
        # 举升电机测试需要传递高度参数
        height = data.get('height')
        result = api_handler.send_command(item_id, ssh_host, ssh_user, ssh_password, height=height)
    elif test_id == 'rotation_motor':
        # 旋转电机测试需要传递角度参数
        angle = data.get('angle')
        result = api_handler.send_command(item_id, ssh_host, ssh_user, ssh_password, angle=angle)
    elif test_id == 'walking_motor':
        # 行走电机测试需要传递距离参数
        distance = data.get('distance')
        result = api_handler.send_command(item_id, ssh_host, ssh_user, ssh_password, distance=distance)
    else:
        # 其他测试类型使用BaseAPI，不传递command_map（允许无命令的测试项）
        result = api_handler.send_command(item_id, command_map=None, ssh_host=ssh_host, ssh_user=ssh_user, ssh_password=ssh_password)
    
    if result.get('status') == 'error':
        return jsonify(result), 400 if '未知' in result.get('message', '') else 500
    return jsonify(result)

@app.route('/api/camera/connect_ssh', methods=['POST'])
def camera_connect_ssh():
    """建立相机测试的SSH连接（进入camera tab时调用）"""
    import paramiko
    
    data = request.json
    ssh_host = data.get('ssh_host')
    ssh_user = data.get('ssh_user') or config.SSH_USER
    ssh_password = data.get('ssh_password') or config.SSH_PASSWORD
    
    if not ssh_host:
        return jsonify({"status": "error", "message": "未提供SSH主机地址"}), 400
    
    with ssh_lock:
        # 如果连接已存在，先关闭旧连接
        if ssh_host in camera_ssh_connections:
            try:
                old_ssh = camera_ssh_connections[ssh_host]['ssh']
                old_ssh.close()
            except:
                pass
            del camera_ssh_connections[ssh_host]
        
        try:
            # 建立新的SSH连接
            ssh = paramiko.SSHClient()
            ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            ssh.connect(
                hostname=ssh_host,
                username=ssh_user,
                password=ssh_password,
                timeout=config.SSH_TIMEOUT,
                look_for_keys=False,
                allow_agent=False
            )
            
            # 保存连接
            camera_ssh_connections[ssh_host] = {
                'ssh': ssh,
                'lock': threading.Lock()
            }
            
            print(f"[相机测试] ✅ SSH连接已建立并保存: {ssh_host}")
            return jsonify({"status": "success", "message": f"SSH连接已建立: {ssh_host}"})
        except Exception as e:
            print(f"[相机测试] ❌ SSH连接失败: {str(e)}")
            return jsonify({"status": "error", "message": f"SSH连接失败: {str(e)}"}), 500

@app.route('/api/camera/disconnect_ssh', methods=['POST'])
def camera_disconnect_ssh():
    """断开相机测试的SSH连接（退出camera tab时调用）"""
    data = request.json
    ssh_host = data.get('ssh_host')
    
    if not ssh_host:
        return jsonify({"status": "error", "message": "未提供SSH主机地址"}), 400
    
    with ssh_lock:
        if ssh_host in camera_ssh_connections:
            try:
                ssh = camera_ssh_connections[ssh_host]['ssh']
                ssh.close()
                print(f"[相机测试] 🔌 SSH连接已断开: {ssh_host}")
            except Exception as e:
                print(f"[相机测试] ⚠️ 断开SSH连接时出错: {str(e)}")
            finally:
                del camera_ssh_connections[ssh_host]
            return jsonify({"status": "success", "message": f"SSH连接已断开: {ssh_host}"})
        else:
            return jsonify({"status": "success", "message": f"SSH连接不存在: {ssh_host}"})

@app.route('/api/check_io', methods=['POST'])
def check_io():
    """检查IO信号状态（根据测试类型路由到对应的API处理器）"""
    data = request.json
    item_id = data.get('item_id')
    test_id = data.get('test_id')  # 可选，用于确定使用哪个API处理器
    ssh_host = data.get('ssh_host')  # SSH主机地址
    ssh_user = data.get('ssh_user')  # SSH用户名
    ssh_password = data.get('ssh_password')  # SSH密码
    vehicle_model = data.get('vehicle_model')  # 车型（用于按键测试的IO映射）
    
    # 如果没有提供test_id，根据item_id推断
    if not test_id:
        if item_id in config.IO_INDEX_MAP:
            test_id = 'light'
        elif vehicle_model:
            # 如果有车型信息，检查是否在按键IO映射中
            button_io_map = config.get_button_io_map(vehicle_model)
            if item_id in button_io_map:
                test_id = 'button'
            else:
                test_id = 'default'
        else:
            # 如果没有车型信息，尝试使用默认X100的映射
            default_button_map = config.get_button_io_map('X100')
            if item_id in default_button_map:
                test_id = 'button'
            else:
                test_id = 'default'
    
    # 获取对应的API处理器
    api_handler = apis.get_api_handler(test_id)
    
    # 调用对应的检查IO方法，传递SSH信息和车型信息
    if test_id in ['button', 'touch', 'display']:
        # 按键测试、触边测试、显示屏测试需要SSH信息、车型信息和测试ID
        result = api_handler.check_io(item_id, ssh_host, ssh_user, ssh_password, vehicle_model, test_id)
    elif test_id == 'light':
        # 灯光测试需要SSH信息
        result = api_handler.check_io(item_id, ssh_host, ssh_user, ssh_password)
    elif test_id == 'camera':
        # 相机测试：检查是否是TOF测试项（使用订阅方式）
        if item_id in ['front_tof', 'rear_tof']:
            # TOF测试：使用订阅方式
            ssh_connection = None
            use_existing_ssh = False
            if ssh_host and ssh_host in camera_ssh_connections:
                ssh_connection = camera_ssh_connections[ssh_host]['ssh']
                use_existing_ssh = True
            result = api_handler.check_tof_subscribe(item_id, ssh_host, ssh_user, ssh_password, vehicle_model, test_id, use_existing_ssh=use_existing_ssh, ssh_connection=ssh_connection)
        else:
            # 其他相机测试：需要IP地址和SSH信息（通过SSH在远程设备上执行ping）
            ip_address = data.get('ip_address')
            # 检查是否有已存在的SSH连接，如果有则传递SSH对象
            ssh_connection = None
            use_existing_ssh = False
            if ssh_host and ssh_host in camera_ssh_connections:
                ssh_connection = camera_ssh_connections[ssh_host]['ssh']
                use_existing_ssh = True
            result = api_handler.check_io(item_id, ssh_host, ssh_user, ssh_password, vehicle_model, test_id, ip_address, use_existing_ssh=use_existing_ssh, ssh_connection=ssh_connection)
    else:
        result = api_handler.check_io(item_id)
    
    if result.get('status') == 'error':
        return jsonify(result), 400 if '未知' in result.get('message', '') else 500
    return jsonify(result)

@app.route('/api/submit_test', methods=['POST'])
def submit_test():
    data = request.json
    # 这里可以保存测试结果到数据库
    print(f"收到测试结果: {data}")
    return jsonify({"status": "success", "message": "测试结果已保存"})

@app.route('/api/test_report', methods=['POST'])
def get_test_report():
    """获取测试报告数据"""
    data = request.json
    test_results = data.get('test_results', {})
    vehicle_model = data.get('vehicle_model', 'X100')
    hostname = data.get('hostname', '')
    test_time = data.get('test_time', '')  # 获取前端传递的测试时间
    
    # 调试：打印接收到的测试结果
    print(f"[报告调试] 接收到的测试结果: {test_results}")
    print(f"[报告调试] 车型: {vehicle_model}")
    print(f"[报告调试] 测试时间: {test_time}")
    
    # 根据车型获取对应的测试类别
    categories = get_categories_by_vehicle(vehicle_model)
    
    # 构建报告数据
    report_data = []
    for category in categories:
        category_id = category['id']
        category_name = category['name']
        category_results = test_results.get(category_id, {})
        
        # 调试：打印每个类别的结果
        print(f"[报告调试] 类别 {category_id} ({category_name}) 的结果: {category_results}")
        
        # 获取该类别的测试详情
        test_detail = test_data.TEST_DETAILS.get(category_id, {})
        sections = test_detail.get('sections', [])
        
        # 根据测试类型过滤测试项
        allowed_item_ids = None
        if category_id == 'camera':
            # 相机测试：只显示CSV配置中指定的设备
            camera_devices = TABdisplayconfig.get_camera_devices_by_device_model(vehicle_model)
            allowed_item_ids = set(camera_devices)
            print(f"[报告调试] 相机测试允许的设备: {allowed_item_ids}")
        elif category_id == 'button':
            # 按键测试：只显示该车型允许的按钮
            allowed_items = config.get_button_test_items(vehicle_model)
            allowed_item_ids = set(allowed_items)
            print(f"[报告调试] 按键测试允许的按钮: {allowed_item_ids}")
        
        for section in sections:
            for item in section.get('items', []):
                item_id = item['id']
                
                # 如果有限制列表，检查是否在允许列表中
                if allowed_item_ids is not None:
                    if item_id not in allowed_item_ids:
                        print(f"[报告调试] 跳过未显示的测试项: {item_id}")
                        continue
                
                item_name = item['name']
                result = category_results.get(item_id, '未测试')
                
                # 调试：打印每个测试项的结果
                print(f"[报告调试] 测试项 {item_id} ({item_name}): {result}")
                
                # 转换结果为中文
                if result == 'normal':
                    result_text = '正常'
                elif result == 'abnormal':
                    result_text = '异常'
                else:
                    result_text = '未测试'
                
                report_data.append({
                    'category': category_name,
                    'item': item_name,
                    'result': result_text
                })
    
    return jsonify({
        'status': 'success',
        'data': report_data,
        'title': '机器人静态测试报告',
        'vehicle_model': vehicle_model,
        'hostname': hostname,
        'test_time': test_time if test_time else '-'  # 使用前端传递的测试时间，如果没有则显示"-"
    })

@app.route('/api/download_report', methods=['POST'])
def download_report():
    """下载测试报告为Excel文件"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        
        data = request.json
        test_results = data.get('test_results', {})
        vehicle_model = data.get('vehicle_model', 'X100')
        hostname = data.get('hostname', '')
        test_time = data.get('test_time', '')  # 获取前端传递的测试时间
        
        # 如果没有传递测试时间，使用"-"
        if not test_time:
            test_time = '-'
        
        # 根据车型获取对应的测试类别
        categories = get_categories_by_vehicle(vehicle_model)
        
        # 创建Excel工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "测试报告"
        
        # 设置标题
        ws.merge_cells('A1:C1')
        title_cell = ws['A1']
        title_cell.value = '机器人静态测试报告'
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 使用前端传递的测试时间（已在上面处理）
        
        # 信息区域样式
        info_label_font = Font(bold=True, size=12)
        info_value_font = Font(size=12)
        
        # 信息区域对齐样式
        info_alignment = Alignment(horizontal='center', vertical='center')
        
        # 型号
        ws['A3'] = '型号:'
        ws['A3'].font = info_label_font
        ws['A3'].alignment = info_alignment
        ws['B3'] = vehicle_model
        ws['B3'].font = info_value_font
        ws['B3'].alignment = info_alignment
        
        # 软件版本（从请求中获取APP_VERSION）
        app_version = data.get('app_version', '-')
        ws['A4'] = '软件版本:'
        ws['A4'].font = info_label_font
        ws['A4'].alignment = info_alignment
        ws['B4'] = app_version if app_version else '-'
        ws['B4'].font = info_value_font
        ws['B4'].alignment = info_alignment
        
        # 设备序列号
        ws['A5'] = '设备序列号:'
        ws['A5'].font = info_label_font
        ws['A5'].alignment = info_alignment
        ws['B5'] = hostname if hostname else '-'
        ws['B5'].font = info_value_font
        ws['B5'].alignment = info_alignment
        
        # 测试时间
        ws['A6'] = '测试时间:'
        ws['A6'].font = info_label_font
        ws['A6'].alignment = info_alignment
        ws['B6'] = test_time
        ws['B6'].font = info_value_font
        ws['B6'].alignment = info_alignment
        
        # 测试人员
        tester = data.get('tester', '张三')
        ws['A7'] = '测试人员:'
        ws['A7'].font = info_label_font
        ws['A7'].alignment = info_alignment
        ws['B7'] = tester
        ws['B7'].font = info_value_font
        ws['B7'].alignment = info_alignment
        
        # 设置表格表头
        headers = ['测试类别', '测试项', '测试结果']
        ws.append([])  # 空行（第8行）
        ws.append(headers)  # 第9行
        
        # 设置表头样式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for col in range(1, 4):
            cell = ws.cell(row=9, column=col)  # 表格表头在第9行
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 填充数据
        row_num = 10  # 数据从第10行开始
        for category in categories:
            category_id = category['id']
            category_name = category['name']
            category_results = test_results.get(category_id, {})
            
            # 获取该类别的测试详情
            test_detail = test_data.TEST_DETAILS.get(category_id, {})
            sections = test_detail.get('sections', [])
            
            # 根据测试类型过滤测试项
            allowed_item_ids = None
            if category_id == 'camera':
                # 相机测试：只显示CSV配置中指定的设备
                camera_devices = TABdisplayconfig.get_camera_devices_by_device_model(vehicle_model)
                allowed_item_ids = set(camera_devices)
            elif category_id == 'button':
                # 按键测试：只显示该车型允许的按钮
                allowed_items = config.get_button_test_items(vehicle_model)
                allowed_item_ids = set(allowed_items)
            
            for section in sections:
                for item in section.get('items', []):
                    item_id = item['id']
                    
                    # 如果有限制列表，检查是否在允许列表中
                    if allowed_item_ids is not None:
                        if item_id not in allowed_item_ids:
                            continue  # 跳过未显示的测试项
                    
                    item_name = item['name']
                    result = category_results.get(item_id, '未测试')
                    
                    # 转换结果为中文
                    if result == 'normal':
                        result_text = '正常'
                    elif result == 'abnormal':
                        result_text = '异常'
                    else:
                        result_text = '未测试'
                    
                    ws.append([category_name, item_name, result_text])
                    
                    # 设置结果列颜色
                    result_cell = ws.cell(row=row_num, column=3)
                    if result_text == '正常':
                        result_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif result_text == '异常':
                        result_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    
                    row_num += 1
        
        # 调整列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        
        # 设置对齐方式
        for row in ws.iter_rows(min_row=5, max_row=row_num-1):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 保存到内存
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 生成文件名，使用测试开始时间戳（如果有），否则使用当前时间
        test_time = data.get('test_time', '')  # 获取前端传递的测试时间
        if test_time and test_time != '-':
            # 将格式从 "2026-01-13 11:13:20" 转换为 "2026-01-13_11-13-20"（文件名中不能有冒号）
            time_str = test_time.replace(' ', '_').replace(':', '-')
        else:
            # 如果没有测试开始时间，使用当前时间
            time_str = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        filename = f'机器人静态测试报告_{vehicle_model}_{time_str}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({"status": "error", "message": f"生成报告失败: {str(e)}"}), 500

@app.route('/api/save_report', methods=['POST'])
def save_report():
    """保存测试报告到本地report文件夹"""
    try:
        import os
        
        # 检查是否有文件
        if 'file' not in request.files:
            return jsonify({"status": "error", "message": "未找到文件"}), 400
        
        file = request.files['file']
        cloudname = request.form.get('cloudname')
        
        if not cloudname:
            return jsonify({"status": "error", "message": "缺少文件名参数"}), 400
        
        # 获取项目根目录
        base_dir = os.path.dirname(os.path.abspath(__file__))
        report_dir = os.path.join(base_dir, 'report')
        
        # 确保report目录存在
        if not os.path.exists(report_dir):
            os.makedirs(report_dir)
            print(f"[保存报告] 创建report目录: {report_dir}")
        
        # 文件完整路径
        file_path = os.path.join(report_dir, cloudname)
        
        # 保存文件
        file.save(file_path)
        
        # 获取文件大小
        cloudsize = os.path.getsize(file_path)
        
        print(f"[保存报告] 文件已保存: {file_path}, 大小: {cloudsize}字节")
        
        return jsonify({
            "status": "success",
            "message": "文件已保存到本地",
            "cloudname": cloudname,
            "cloudsize": cloudsize,
            "file_path": file_path
        })
    except Exception as e:
        return jsonify({"status": "error", "message": f"保存文件失败: {str(e)}"}), 500

@app.route('/api/upload_to_cloud', methods=['POST'])
def upload_to_cloud():
    """上传测试报告到飞书云文档"""
    try:
        import os
        import subprocess
        
        data = request.json
        cloudname = data.get('cloudname')  # 文件名（含后缀）
        cloudsize = data.get('cloudsize')  # 文件大小（字节）
        
        if not cloudname or not cloudsize:
            return jsonify({"status": "error", "message": "缺少文件名或文件大小参数"}), 400
        
        # 获取项目根目录
        base_dir = os.path.dirname(os.path.abspath(__file__))
        report_dir = os.path.join(base_dir, 'report')
        
        # 文件完整路径
        file_path = os.path.join(report_dir, cloudname)
        
        # 检查文件是否存在
        if not os.path.exists(file_path):
            return jsonify({"status": "error", "message": f"文件不存在: {file_path}"}), 400
        
        # 验证文件大小
        actual_size = os.path.getsize(file_path)
        if actual_size != cloudsize:
            print(f"[云端同步] ⚠️ 文件大小不匹配: 预期{cloudsize}字节，实际{actual_size}字节，使用实际大小")
            cloudsize = actual_size
        
        print(f"[云端同步] 准备上传文件: {cloudname}, 大小: {cloudsize}字节, 路径: {file_path}")
        
        # 执行上传脚本
        script_path = os.path.join(base_dir, 'fs_files_upload.py')
        if not os.path.exists(script_path):
            return jsonify({"status": "error", "message": f"上传脚本不存在: {script_path}"}), 500
        
        # 调用Python脚本，传递参数
        try:
            result = subprocess.run(
                ['python', script_path, cloudname, str(cloudsize)],
                capture_output=True,
                text=True,
                timeout=60,
                cwd=base_dir
            )
            
            print(f"[云端同步] 脚本执行完成，返回码: {result.returncode}")
            print(f"[云端同步] 脚本输出:\n{result.stdout}")
            if result.stderr:
                print(f"[云端同步] 脚本错误输出:\n{result.stderr}")
            
            if result.returncode == 0:
                return jsonify({
                    "status": "success",
                    "message": f"文件已成功上传到云端: {cloudname}",
                    "output": result.stdout
                })
            else:
                return jsonify({
                    "status": "error",
                    "message": f"上传脚本执行失败: {result.stderr or result.stdout}",
                    "output": result.stdout,
                    "error": result.stderr
                }), 500
                
        except subprocess.TimeoutExpired:
            return jsonify({"status": "error", "message": "上传脚本执行超时"}), 500
        except Exception as e:
            return jsonify({"status": "error", "message": f"执行上传脚本失败: {str(e)}"}), 500
        print(f"[云端同步] 文件大小: {len(file_content)} 字节")
        print(f"[云端同步] 目标文件夹: https://thundersoft.feishu.cn/drive/folder/K5RAfwXaNl0Mc2dmp5dcpmFinKb")
        
        # 临时方案：保存文件到本地，提示用户手动上传
        # 或者返回文件夹链接，让用户手动上传
        
        return jsonify({
            "status": "success",
            "message": f"报告已准备上传\n文件名: {filename}\n请访问以下链接手动上传:\nhttps://thundersoft.feishu.cn/drive/folder/K5RAfwXaNl0Mc2dmp5dcpmFinKb"
        })
        
    except Exception as e:
        print(f"[云端同步] 上传失败: {str(e)}")
        return jsonify({"status": "error", "message": f"上传失败: {str(e)}"}), 500

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
